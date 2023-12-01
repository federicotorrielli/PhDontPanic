import argparse

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Set up command-line argument parsing
parser = argparse.ArgumentParser(description="Compare two CSV files.")
parser.add_argument("csv1", type=str, help="Path to the first CSV file.")
parser.add_argument("csv2", type=str, help="Path to the second CSV file.")
parser.add_argument("xlsx", type=str, help="Path to the output Excel file.")
parser.add_argument(
    "green_rows", type=int, help="Number of unmarked rows to highlight in green."
)
args = parser.parse_args()

# Load CSV files into pandas dataframes
df1 = pd.read_csv(args.csv1)
df2 = pd.read_csv(args.csv2)

# Reorder rows based on 'Title' column
df1 = df1.sort_values(by="Title")
df2 = df2.sort_values(by="Title")

# Find the rows that are in df1 but not in df2
df_diff = df1.loc[~df1["Title"].isin(df2["Title"])]

# Select only the desired columns
columns = [
    "Item Type",
    "Publication Year",
    "Author",
    "Title",
    "Extra",
    "Abstract Note",
    "Publication Title",
]
df1 = df1[columns]
df2 = df2[columns]
df_diff = df_diff[columns]

# Write the dataframes to an Excel file
with pd.ExcelWriter(args.xlsx) as writer:
    df1.to_excel(writer, sheet_name="Original", index=False)
    df2.to_excel(writer, sheet_name="In Progress", index=False)
    df_diff.to_excel(writer, sheet_name="Removed", index=False)

# Load the Excel file with openpyxl
wb = load_workbook(filename=args.xlsx)

# Define a green fill style
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Get the worksheet for the second CSV file
ws = wb["In Progress"]

# Highlight the first n rows as green
# Mark rows containing "dubbio" in the "Extra" column as yellow
for row in ws.iter_rows(min_row=2, max_row=args.green_rows + 1):
    for cell in row:
        cell.fill = green_fill
    if "dubbio" in row[4].value:
        for cell in row:
            cell.fill = yellow_fill

# Save the Excel file
wb.save(filename=args.xlsx)
